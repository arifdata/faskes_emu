from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, FileResponse
from collections import OrderedDict, Counter
import operator
from socket import gethostname, gethostbyname

# Create your views here.
def contact_developer(request):
    return render(request, 'laporan/hubungi_developer.html')

def all_stock(request):
    bmhp = {}
    from apotek.models import StokObatApotek, StokObatGudang
    q = StokObatGudang.objects.all()
    for i in q:
        bmhp[i.nama_obat.nama_obat] = i.jml

    r = StokObatApotek.objects.all()
    for i in r:
        bmhp[i.nama_obat.nama_obat] += i.jml

    ctx = {
        'bmhp': bmhp
    }
        
    return render(request, 'laporan/all_stok.html', ctx)
    
def index_page(request):
    import datetime
    from statistics import mean
    from poli.models import DataKunjungan
    from apotek.models import Resep, Pengeluaran, Penerimaan
    # get tanggal sekarang
    now = datetime.datetime.now()
    three_month_before = now - datetime.timedelta(days=90)
    three_month_after = now + datetime.timedelta(days=90)

    #dict utk menampung data sementara
    raw_data_kunjungan = {}
    raw_data_penyakit = {}
    raw_data_obat = {}
    raw_data_penulis = {}
    raw_data_ed = {}

    ed = Penerimaan.objects.select_related('nama_barang').filter(tgl_kadaluarsa__gte=f"{three_month_before.year}-{three_month_before.month}-{three_month_before.day}", tgl_kadaluarsa__lte=f"{three_month_after.year}-{three_month_after.month}-{three_month_after.day}").iterator()

    for i in ed:
        raw_data_ed[i.nama_barang.nama_obat] = [f"{i.tgl_kadaluarsa.year}-{i.tgl_kadaluarsa.month}-{i.tgl_kadaluarsa.day}", f"{(i.tgl_kadaluarsa + datetime.timedelta(days=2)).year}-{(i.tgl_kadaluarsa + datetime.timedelta(days=2)).month}-{(i.tgl_kadaluarsa + datetime.timedelta(days=2)).day}"]

    #query data dari awal bulan sekarang sampai sekarang
    query = DataKunjungan.objects.select_related('penulis_resep').filter(tgl_kunjungan__gte="{}-{}-1".format(now.year, now.month), tgl_kunjungan__lte="{}-{}-{}".format(now.year, now.month, now.day))

    # hitung penulis resep
    for data in query:
        if data.penulis_resep.nama_peresep not in raw_data_penulis:
            raw_data_penulis[data.penulis_resep.nama_peresep] = 1
        else:
            raw_data_penulis[data.penulis_resep.nama_peresep] += 1

    # memasukkan diagnosa ke raw_data_penyakit dan menghitung akumulasinya
    for data in query:
        for diag in data.diagnosa.values('diagnosa'):
            if diag['diagnosa'] not in raw_data_penyakit:
                raw_data_penyakit[diag['diagnosa']] = 1
            else:
                raw_data_penyakit[diag['diagnosa']] += 1
    
    # memasukkan tanggal beserta jumlah kunjungannya respectively ke dict raw_data_kunjungan
    for data in query:
        b = DataKunjungan.objects.filter(tgl_kunjungan=data.tgl_kunjungan).count()
        raw_data_kunjungan[data.tgl_kunjungan.strftime('%d-%m')] = b

    # mengurutkan data sesuai urutan tanggal
    cleaned_data_kunjungan = OrderedDict(sorted(raw_data_kunjungan.items()))

    # mengurutkan data sesuai urutan tanggal
    cleaned_data_penyakit = OrderedDict(sorted(raw_data_penyakit.items()))

    query_obat = Resep.objects.select_related('nama_obat__nama_obat').filter(kunjungan_pasien__tgl_kunjungan__gte="{}-{}-1".format(now.year, now.month), kunjungan_pasien__tgl_kunjungan__lte="{}-{}-{}".format(now.year, now.month, now.day)).iterator()

    # Handling perhitungan jumlah obat terpakai
    for data in query_obat:
        if data.nama_obat.nama_obat.nama_obat not in raw_data_obat:
            raw_data_obat[data.nama_obat.nama_obat.nama_obat] = data.jumlah
        else:
            raw_data_obat[data.nama_obat.nama_obat.nama_obat] += data.jumlah

    # urutkan dictionary obat berdasarkan jumlah terbanyak
    cleaned_data_obat = dict(sorted(raw_data_obat.items(), key=operator.itemgetter(1),reverse=True))
    
    rerata = 0
    if len(list(cleaned_data_kunjungan.values())) > 0:
        rerata = mean(list(cleaned_data_kunjungan.values()))

    addr = gethostbyname(gethostname())
    if addr == "127.0.0.1":
        addr = ""
        
    penyakit = [{'x':k, 'y':v} for k, v in cleaned_data_penyakit.items()]
    maksimum = 0 if len(cleaned_data_kunjungan.values()) == 0 else max(list(cleaned_data_kunjungan.values()))

    context = {
        'bln': now.strftime("%B"), 
        'thn': now.year, 
        'labels_kunjungan': list(cleaned_data_kunjungan.keys()), 
        'data_kunjungan': list(cleaned_data_kunjungan.values()),
        'rerata_kunjungan': rerata,
        'maks': maksimum,
        'labels_obat_terbanyak': list(cleaned_data_obat.keys())[0:20],
        'data_obat_terbanyak': list(cleaned_data_obat.values())[0:20],
        'addr': addr,
        'labels_penulis_resep': list(raw_data_penulis.keys()),
        'data_penulis_resep': list(raw_data_penulis.values()),
        'penyakit': penyakit,
        'ed': raw_data_ed,
    }

    return render(request, 'laporan/index.html', context)

def laporan_page(request):
    return render(request, 'laporan/laporan_generator.html')

@login_required
def penggunaan_harian(request):
    from apotek.models import Resep
    from .forms import TglForm
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = TglForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            raw_data_bmhp_apt = {}
            
            q = Resep.objects.select_related('nama_obat__nama_obat').filter(kunjungan_pasien__tgl_kunjungan__gte=request.POST.get("tanggal1"), kunjungan_pasien__tgl_kunjungan__lte=request.POST.get("tanggal2")).iterator()

            for data in q:
                #print(data.kunjungan_pasien.tgl_kunjungan, data.nama_obat.nama_obat, data.jumlah, data.kunjungan_pasien.nama_pasien.nama_pasien)
                if str(data.kunjungan_pasien.tgl_kunjungan) not in raw_data_bmhp_apt:
                    raw_data_bmhp_apt[str(data.kunjungan_pasien.tgl_kunjungan)] = {data.nama_obat.nama_obat.nama_obat: data.jumlah}                    
                elif data.nama_obat.nama_obat.nama_obat in raw_data_bmhp_apt[str(data.kunjungan_pasien.tgl_kunjungan)]:
                    raw_data_bmhp_apt[str(data.kunjungan_pasien.tgl_kunjungan)][data.nama_obat.nama_obat.nama_obat] += data.jumlah
                elif data.nama_obat.nama_obat.nama_obat not in raw_data_bmhp_apt[str(data.kunjungan_pasien.tgl_kunjungan)]:
                    raw_data_bmhp_apt[str(data.kunjungan_pasien.tgl_kunjungan)][data.nama_obat.nama_obat.nama_obat] = data.jumlah
                else:
                    pass
            
            context = {
                'startdate': request.POST.get("tanggal1"),
                'enddate': request.POST.get("tanggal2"),
                'val': raw_data_bmhp_apt,
            }
            return render(request, 'laporan/penggunaan_harian.html', context)

    # if a GET (or any other method) we'll create a blank form
    else:
        form = TglForm()

    return render(request, 'laporan/form_penggunaan_harian.html', {'form': form})

@login_required
def cetak_so_gudang(request):
    from apotek.models import SOGudang
    from .forms import SOGudangForm

    if request.method == 'POST':
        form = SOGudangForm(request.POST)

        if form.is_valid():
            dataso = SOGudang.objects.get(pk=request.POST.get("pilihan"))
            context = {
                'tanggal': dataso.tgl,
                'data': dataso.data,
            }
            return render(request, 'laporan/hasil_so_gudang.html', context)

    else:
        form = SOGudangForm()

    return render(request, 'laporan/form_cetak_so_gudang.html', {'form': form})

@login_required
def cetak_so_apotek(request):
    from apotek.models import SOApotek
    from .forms import SOApotekForm

    if request.method == 'POST':
        form = SOApotekForm(request.POST)

        if form.is_valid():
            dataso = SOApotek.objects.get(pk=request.POST.get("pilihan"))
            context = {
                'tanggal': dataso.tgl,
                'data': dataso.data,
            }
            return render(request, 'laporan/hasil_so_apotek.html', context)

    else:
        form = SOApotekForm()

    return render(request, 'laporan/form_cetak_so_apotek.html', {'form': form})


@login_required
def penggunaan_bmhp(request):
    from apotek.models import Resep, Pengeluaran
    from .forms import TglForm
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = TglForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            raw_data_bmhp_apt = {}
            raw_data_bmhp_unit = {}
            kunci = []
            cleaned_data_bmhp_unit = {}
            
            q = Resep.objects.select_related('nama_obat__nama_obat').filter(kunjungan_pasien__tgl_kunjungan__gte=request.POST.get("tanggal1"), kunjungan_pasien__tgl_kunjungan__lte=request.POST.get("tanggal2")).iterator()
            q2 = Pengeluaran.objects.select_related('nama_barang__nama_obat').filter(keluar_barang__tgl_keluar__gte=request.POST.get("tanggal1"), keluar_barang__tgl_keluar__lte=request.POST.get("tanggal2")).exclude(keluar_barang__tujuan__nama="APOTEK").iterator()

            for data in q:
                if data.nama_obat.nama_obat.nama_obat not in raw_data_bmhp_apt:
                    raw_data_bmhp_apt[data.nama_obat.nama_obat.nama_obat] = data.jumlah
                else:
                    raw_data_bmhp_apt[data.nama_obat.nama_obat.nama_obat] += data.jumlah

            cleaned_data_obat = dict(sorted(raw_data_bmhp_apt.items(), key=operator.itemgetter(1),reverse=True))
            cleaned_data_obat = OrderedDict(sorted(cleaned_data_obat.items()))

            for data in q2:
                a = {}
                a[data.nama_barang.nama_obat.nama_obat] = data.jumlah
                if data.keluar_barang.tujuan.nama not in raw_data_bmhp_unit.keys():
                    raw_data_bmhp_unit[data.keluar_barang.tujuan.nama] = [a]
                    kunci.append(data.keluar_barang.tujuan.nama)
                else:                    
                    raw_data_bmhp_unit[data.keluar_barang.tujuan.nama].append(a)

            for d in kunci:
                c = Counter()
                for x in raw_data_bmhp_unit[d]:
                    c.update(x)
                    cleaned_data_bmhp_unit[d] = dict(c)
            
            context = {
                'startdate': request.POST.get("tanggal1"),
                'enddate': request.POST.get("tanggal2"),
                'val': cleaned_data_obat,
                'unit': cleaned_data_bmhp_unit,
            }
            return render(request, 'laporan/penggunaan_bmhp.html', context)

    # if a GET (or any other method) we'll create a blank form
    else:
        form = TglForm()

    return render(request, 'laporan/form_penggunaan_bmhp.html', {'form': form})

def genKunjChart(x, y):
    import pygal
    from pygal.style import LightenStyle
    
    lineChart = pygal.Line(fill=True, style=LightenStyle('#047857'), print_values=True, print_values_position="top", x_title="Tanggal", y_title="Jml Lembar Resep", show_legend=False, x_label_rotation=20, show_x_guides=True, margin=20)
    lineChart.x_labels = x
    lineChart.add("Jml Lembar Resep", y)
    return lineChart.render_data_uri()

def genObatTerbanyakChart(label, data):
    import pygal
    from pygal.style import SolidColorStyle

    barChart = pygal.HorizontalBar(title="Penggunaan 20 Obat Terbanyak di Apotek", print_values=True, legend_at_bottom=True, legend_box_size=10, legend_at_bottom_columns=2, style=SolidColorStyle)
    for i in range(0, len(label)):
        barChart.add(label[i], data[i])
    return barChart.render_data_uri()

def genPeresepChart(label, data):
    import pygal
    from pygal.style import SolidColorStyle

    total = sum(data)
    pieChart = pygal.Pie(title="Persentase Lembar Resep", print_values=True, style=SolidColorStyle, legend_at_bottom=True, margin=10, print_labels=True)
    for i in range(0, len(label)):
        persentase = (data[i] / total) * 100
        pieChart.add(label[i], [{'value': float("{:.2f}".format(persentase)), 'label': label[i]}])
    pieChart.value_formatter = lambda x: "{} %".format(x)
    return pieChart.render_data_uri()

def genTenPenyakit(data):
    import pygal
    from pygal.style import SolidColorStyle
    labels = []
    for key in data.keys():
        labels.append(key)

    tm = pygal.Treemap(print_labels=True, print_values=True, legend_at_bottom=True, legend_at_bottom_columns=1, margin=10, style=SolidColorStyle)
    for i in range(0, len(labels)):
        tm.add(labels[i], [{'value': data[labels[i]], 'label': 'Essential (primary) hypertension' if labels[i] == 'Essential (primary) hypertension (I10)' else labels[i].split(' (')[0]}])
    return tm.render_data_uri()

@login_required
def laporan_semua(request):
    import datetime
    from statistics import mean
    from poli.models import DataKunjungan
    from apotek.models import Resep, Pengeluaran, Penerimaan, KartuStokApotek, KartuStokGudang, StokObatGudang, StokObatApotek
    from .forms import TglForm
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = TglForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
        
            now = datetime.datetime.now()
            three_month_before = now - datetime.timedelta(days=90)
            three_month_after = now + datetime.timedelta(days=90)

            #dict utk menampung data sementara
            raw_data_kunjungan = {}
            raw_data_penyakit = {}
            raw_penyakit_per_lokasi = {}
            raw_data_obat = {}
            raw_data_penulis = {}
            raw_data_ed = {}
            raw_data_bmhp_unit = {}
            kunci = []
            cleaned_data_bmhp_unit = {}
            raw_data_bmhp_apt = {}
            raw_data_penerimaan = {}
            raw_data_gd_ke_apt = {}
            kunci2 = []
            cleaned_data_gd_ke_apt = {}

            ed = Penerimaan.objects.select_related('nama_barang').filter(tgl_kadaluarsa__gte=f"{three_month_before.year}-{three_month_before.month}-{three_month_before.day}", tgl_kadaluarsa__lte=f"{three_month_after.year}-{three_month_after.month}-{three_month_after.day}").iterator()

            for i in ed:
                raw_data_ed[i.nama_barang.nama_obat] = [f"{i.tgl_kadaluarsa.year}-{i.tgl_kadaluarsa.month}-{i.tgl_kadaluarsa.day}", f"{(i.tgl_kadaluarsa + datetime.timedelta(days=2)).year}-{(i.tgl_kadaluarsa + datetime.timedelta(days=2)).month}-{(i.tgl_kadaluarsa + datetime.timedelta(days=2)).day}"]

            #query data dari awal bulan sekarang sampai sekarang
            query = DataKunjungan.objects.select_related('penulis_resep').filter(tgl_kunjungan__gte=request.POST.get("tanggal1"), tgl_kunjungan__lte=request.POST.get("tanggal2"))

            # query alamat
            for data in query:
               if data.nama_pasien.alamat not in raw_penyakit_per_lokasi.keys():
                raw_penyakit_per_lokasi[data.nama_pasien.alamat] = {}
                for penyakit in data.diagnosa.values('diagnosa'):
                    if penyakit['diagnosa'] not in raw_penyakit_per_lokasi[data.nama_pasien.alamat].keys():
                        raw_penyakit_per_lokasi[data.nama_pasien.alamat].update({penyakit['diagnosa']: 1})
                    else:
                        raw_penyakit_per_lokasi[data.nama_pasien.alamat][penyakit['diagnosa']] += 1
               else:
                   for penyakit in data.diagnosa.values('diagnosa'):
                       if penyakit['diagnosa'] not in raw_penyakit_per_lokasi[data.nama_pasien.alamat].keys():
                           raw_penyakit_per_lokasi[data.nama_pasien.alamat].update({penyakit['diagnosa']: 1})
                       else:
                            raw_penyakit_per_lokasi[data.nama_pasien.alamat][penyakit['diagnosa']] += 1
            #print(raw_penyakit_per_lokasi)

            # hitung penulis resep
            for data in query:
                if data.penulis_resep.nama_peresep not in raw_data_penulis:
                    raw_data_penulis[data.penulis_resep.nama_peresep] = 1
                else:
                    raw_data_penulis[data.penulis_resep.nama_peresep] += 1

            # memasukkan diagnosa ke raw_data_penyakit dan menghitung akumulasinya
            for data in query:
                for diag in data.diagnosa.values('diagnosa'):
                    if diag['diagnosa'] not in raw_data_penyakit:
                        raw_data_penyakit[diag['diagnosa']] = 1
                    else:
                        raw_data_penyakit[diag['diagnosa']] += 1
            #print(raw_data_penyakit)
            
            # memasukkan tanggal beserta jumlah kunjungannya respectively ke dict raw_data_kunjungan
            for data in query:
                b = DataKunjungan.objects.filter(tgl_kunjungan=data.tgl_kunjungan).count()
                raw_data_kunjungan[data.tgl_kunjungan.strftime('%d-%m')] = b

            q = Resep.objects.select_related('nama_obat__nama_obat').filter(kunjungan_pasien__tgl_kunjungan__gte=request.POST.get("tanggal1"), kunjungan_pasien__tgl_kunjungan__lte=request.POST.get("tanggal2")).iterator()

            for data in q:
                if data.nama_obat.nama_obat.nama_obat not in raw_data_bmhp_apt:
                    raw_data_bmhp_apt[data.nama_obat.nama_obat.nama_obat] = data.jumlah
                else:
                    raw_data_bmhp_apt[data.nama_obat.nama_obat.nama_obat] += data.jumlah

            cleaned_data_apt = dict(sorted(raw_data_bmhp_apt.items(), key=operator.itemgetter(1),reverse=True))
            cleaned_data_apt = OrderedDict(sorted(cleaned_data_apt.items()))
            
            q2 = Pengeluaran.objects.select_related('nama_barang__nama_obat').filter(keluar_barang__tgl_keluar__gte=request.POST.get("tanggal1"), keluar_barang__tgl_keluar__lte=request.POST.get("tanggal2")).exclude(keluar_barang__tujuan__nama="APOTEK").iterator()
            q3 = Pengeluaran.objects.select_related('nama_barang__nama_obat').filter(keluar_barang__tgl_keluar__gte=request.POST.get("tanggal1"), keluar_barang__tgl_keluar__lte=request.POST.get("tanggal2")).filter(keluar_barang__tujuan__nama="APOTEK").iterator()

            for data in q2:
                a = {}
                a[data.nama_barang.nama_obat.nama_obat] = data.jumlah
                if data.keluar_barang.tujuan.nama not in raw_data_bmhp_unit.keys():
                    raw_data_bmhp_unit[data.keluar_barang.tujuan.nama] = [a]
                    kunci.append(data.keluar_barang.tujuan.nama)
                else:                    
                    raw_data_bmhp_unit[data.keluar_barang.tujuan.nama].append(a)

            for d in kunci:
                c = Counter()
                for x in raw_data_bmhp_unit[d]:
                    c.update(x)
                    cleaned_data_bmhp_unit[d] = dict(c)

            for data in q3:
                a = {}
                a[data.nama_barang.nama_obat.nama_obat] = data.jumlah
                if data.keluar_barang.tujuan.nama not in raw_data_gd_ke_apt.keys():
                    raw_data_gd_ke_apt[data.keluar_barang.tujuan.nama] = [a]
                    kunci2.append(data.keluar_barang.tujuan.nama)
                else:                    
                    raw_data_gd_ke_apt[data.keluar_barang.tujuan.nama].append(a)

            for d in kunci2:
                c = Counter()
                for x in raw_data_gd_ke_apt[d]:
                    c.update(x)
                    cleaned_data_gd_ke_apt[d] = dict(c)

            #print(cleaned_data_gd_ke_apt["APOTEK"])

            terima = Penerimaan.objects.filter(terima_barang__tgl_terima__gte=request.POST.get("tanggal1")).filter(terima_barang__tgl_terima__lte=request.POST.get("tanggal2")).order_by("terima_barang__tgl_terima")
            for data in terima:
                raw_data_penerimaan[data.id] = [data.nama_barang.nama_obat, data.terima_barang.tgl_terima, data.terima_barang.sumber.nama, data.no_batch, data.tgl_kadaluarsa, data.jumlah]
                #print(x.nama_barang.nama_obat, x.jumlah, x.terima_barang.tgl_terima, x.id)
                
            # mengurutkan data sesuai urutan tanggal
            cleaned_data_kunjungan = OrderedDict(sorted(raw_data_kunjungan.items()))

            # mengurutkan data sesuai urutan tanggal
            cleaned_data_penyakit = OrderedDict(sorted(raw_data_penyakit.items()))

            query_obat = Resep.objects.select_related('nama_obat__nama_obat').filter(kunjungan_pasien__tgl_kunjungan__gte=request.POST.get("tanggal1"), kunjungan_pasien__tgl_kunjungan__lte=request.POST.get("tanggal2")).iterator()

            # Handling perhitungan jumlah obat terpakai
            for data in query_obat:
                if data.nama_obat.nama_obat.nama_obat not in raw_data_obat:
                    raw_data_obat[data.nama_obat.nama_obat.nama_obat] = data.jumlah
                else:
                    raw_data_obat[data.nama_obat.nama_obat.nama_obat] += data.jumlah

            # urutkan dictionary obat berdasarkan jumlah terbanyak
            cleaned_data_obat = dict(sorted(raw_data_obat.items(), key=operator.itemgetter(1),reverse=True))
            
            rerata = 0
            if len(list(cleaned_data_kunjungan.values())) > 0:
                rerata = mean(list(cleaned_data_kunjungan.values()))
                
            penyakit = [{'x':k, 'y':v} for k, v in cleaned_data_penyakit.items()]
            maksimum = 0 if len(cleaned_data_kunjungan.values()) == 0 else max(list(cleaned_data_kunjungan.values()))
            minimum = 0 if len(cleaned_data_kunjungan.values()) == 0 else min(list(cleaned_data_kunjungan.values()))

            #Get stok apotek
            nama_bmhp_apotek = []
            q_nama_bmhp_apotek = StokObatApotek.objects.all().iterator()
            for i in q_nama_bmhp_apotek:
                nama_bmhp_apotek.append(i.nama_obat.nama_obat)

            raw_data_stok_apotek = {}
            for bmhp in nama_bmhp_apotek:
                try:
                    stok_akhir = KartuStokApotek.objects.filter(nama_obat__nama_obat=bmhp).filter(tgl__gte=request.POST.get("tanggal1"), tgl__lte=request.POST.get("tanggal2")).last()
                    #print(stok_akhir.nama_obat.nama_obat, stok_akhir.sisa_stok)
                    raw_data_stok_apotek[stok_akhir.nama_obat.nama_obat] = stok_akhir.sisa_stok
                except:
                    stok_akhir = KartuStokApotek.objects.filter(nama_obat__nama_obat=bmhp).last()
                    #print(stok_akhir.nama_obat.nama_obat, stok_akhir.sisa_stok)
                    raw_data_stok_apotek[stok_akhir.nama_obat.nama_obat] = stok_akhir.sisa_stok

            # Get stok gudang
            nama_bmhp_gudang = []
            q_nama_bmhp_gudang = StokObatGudang.objects.all().iterator()
            for i in q_nama_bmhp_gudang:
                if i.nama_obat.nama_obat == "zzdummydata":
                    pass
                else:
                    nama_bmhp_gudang.append(i.nama_obat.nama_obat)

            raw_data_stok_gudang = {}

            for i in nama_bmhp_gudang:
                sa = KartuStokGudang.objects.filter(nama_obat__nama_obat=i).filter(tgl__gte=request.POST.get("tanggal1"), tgl__lte=request.POST.get("tanggal2")).last()
                if sa is None:
                    sa = KartuStokGudang.objects.filter(nama_obat__nama_obat=i).last()
                    raw_data_stok_gudang[sa.nama_obat.nama_obat] = sa.sisa_stok
                else:
                    raw_data_stok_gudang[sa.nama_obat.nama_obat] = sa.sisa_stok

            context = {
                #'startdate': request.POST.get("tanggal1"),
                'startdate': datetime.datetime.strptime(request.POST.get("tanggal1"), "%Y-%m-%d").date(),
                'sdate': request.POST.get("tanggal1")[5:7],
                'edate': request.POST.get("tanggal1")[0:4],
                #'enddate': request.POST.get("tanggal2"),
                'enddate': datetime.datetime.strptime(request.POST.get("tanggal2"), "%Y-%m-%d").date(),
                'val': query,
                #'labels_kunjungan': list(cleaned_data_kunjungan.keys()),
                'kunChart64': genKunjChart([x[0:2] for x in cleaned_data_kunjungan.keys()], list(cleaned_data_kunjungan.values())),
                'terbanyak': maksimum,
                'tersedikit': minimum,
                'total_resep': sum(list(cleaned_data_kunjungan.values())),
                'rerata_kunjungan': rerata,
                'maks': maksimum,
                'peresepanChart64': genObatTerbanyakChart(list(cleaned_data_obat.keys())[0:20], list(cleaned_data_obat.values())[0:20]),
                'peresepChart64': genPeresepChart(list(raw_data_penulis.keys()), list(raw_data_penulis.values())),
                'peresep': raw_data_penulis,
                'toppenyakit': genTenPenyakit(OrderedDict(sorted(cleaned_data_penyakit.items(), reverse=True, key=operator.itemgetter(1))[0:10])),
                'penyakit': OrderedDict(sorted(cleaned_data_penyakit.items(), reverse=True, key=operator.itemgetter(1))),
                'lokasi_penyakit': raw_penyakit_per_lokasi,
                'penerimaan': raw_data_penerimaan,
                #'penyakit': sorted(penyakit, key=lambda k: k['y'], reverse=True)[0:1],
                'ed': raw_data_ed,
                'unit': cleaned_data_bmhp_unit,
                'gd_ke_apt': cleaned_data_gd_ke_apt["APOTEK"],
                'apt': cleaned_data_apt,
                'stok_apt': OrderedDict(sorted(raw_data_stok_apotek.items())),
                'stok_gdg': OrderedDict(sorted(raw_data_stok_gudang.items())),
            }
            #print(OrderedDict(sorted(cleaned_data_penyakit.items(), reverse=True, key=operator.itemgetter(1))))
            return render(request, 'laporan/laporan_semua.html', context)

    # if a GET (or any other method) we'll create a blank form
    else:
        form = TglForm()

    return render(request, 'laporan/form_laporan_semua.html', {'form': form})

def pilih_kartu(kartu, tgl_awal, tgl_akhir):
    from apotek.models import KartuStokApotek, KartuStokGudang
    if kartu == "apotek":
        q = KartuStokApotek.objects.select_related('nama_obat').filter(tgl__gte=tgl_awal, tgl__lte=tgl_akhir)
    else:
        q = KartuStokGudang.objects.select_related('nama_obat').filter(tgl__gte=tgl_awal, tgl__lte=tgl_akhir)
    item_set, raw_data = [], {}
    
    for x in q:
        item_set.append(x.nama_obat.nama_obat)
    item_set = list(set(item_set))

    for item in item_set:
        raw_data[item] = []
        data = q.filter(nama_obat__nama_obat=item)
        for x in data:
            raw_data[item].append([x.tgl.strftime("%Y-%m-%d"), x.unit.title(), x.stok_terima, x.stok_keluar, x.sisa_stok, x.ket])

    #pprint.pprint(raw_data)
    return raw_data

@login_required
def cetak_kartu_stok(request):
    from .forms import TglChoiceForm
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = TglChoiceForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            from openpyxl.workbook import Workbook
            from openpyxl.styles import Border, Font, Side, Alignment
            from tempfile import NamedTemporaryFile
            
            raw_data = pilih_kartu(request.POST.get("pilihan"), request.POST.get("tanggal1"), request.POST.get("tanggal2"))
            ft = Font(name='Calibri', size=6.5)
            brd = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
            
            wb = Workbook()
            del wb["Sheet"]
            try:
                for obat, data in raw_data.items():
                    no_char = "/\[]*?:"
                    for c in no_char:
                        if c in obat:
                            obat = obat.replace(c, " ")
                        else:
                            pass
                    ws = wb.create_sheet(obat)
                    ws.set_printer_settings(5, ws.ORIENTATION_LANDSCAPE)
                    ws.page_margins.left = 0.2
                    ws.page_margins.right = 0.2
                    ws.page_margins.bottom = 0.2
                    ws.page_margins.top = 0.2

                    for idx, rows in enumerate(data):
                        if idx % 35 == 0:
                            #ws.append(["KARTU STOK{}".format(90 * " ")])
                            ws.append(["KARTU STOK"])
                            ws.append(["Item : {}".format(obat.title())])
                            ws.append(["Lokasi : {}".format(request.POST.get("pilihan").title())])
                            ws.append(["{}".format(110 * " ")])
                            ws.append(["Tanggal", "Unit|Batch", "Terima", "Keluar", "Sisa", "Ket/ED"])
                            ws.append(rows)
                        else:
                            ws.append(rows)

                    for cells in ws.rows:
                        for cell in cells:
                            if cell.value == "KARTU STOK" or str(cell.value).startswith("Item") or str(cell.value).startswith("Lokasi"):
                                cell.border = brd
                                cell.font = Font(name='Calibri', size=7, bold=True)
                                cell.alignment = Alignment(horizontal='center')
                                ws.merge_cells('A{}:F{}'.format(cell.row, cell.row))
                            elif cell.value == "Tanggal" or cell.value == "Unit|Batch" or cell.value == "Terima" or cell.value == "Keluar" or cell.value == "Sisa" or cell.value == "Ket/ED":
                                cell.font = Font(name='Calibri', size=7, bold=True)
                                cell.border = brd
                                cell.alignment = Alignment(horizontal='center')
                            else:
                                cell.font = ft
                                cell.border = brd
                            
                    ws.column_dimensions['A'].width = 7
                    ws.column_dimensions['B'].width = 9
                    ws.column_dimensions['C'].width = 4
                    ws.column_dimensions['D'].width = 4
                    ws.column_dimensions['E'].width = 4
                    ws.column_dimensions['F'].width = 6

                tmp = NamedTemporaryFile(delete=False)
                with open(tmp.name) as fi:
                    wb.save(tmp.name)
                    return FileResponse(open(tmp.name, 'rb'), as_attachment=True, filename="kartu_stok_{}_{}_to_{}.xlsx".format(request.POST.get("pilihan"), request.POST.get("tanggal1"), request.POST.get("tanggal2")))

            except IndexError:
                ctx = {
                    'gagal': "Tidak ada data. Coba tanggal lain!",
                }
                ctx.update({'form': form})
                #return return render(request, 'laporan/form_cetak_kartu.html', {'form': form})
                return render(request, 'laporan/form_cetak_kartu.html', context=ctx)


    # if a GET (or any other method) we'll create a blank form
    else:
        form = TglChoiceForm()

    return render(request, 'laporan/form_cetak_kartu.html', {'form': form})

@login_required
def lap_narko_psiko(request):
    from .forms import TglForm
    import json
    if request.method == 'POST':
        form = TglForm(request.POST)
        if form.is_valid():
            raw_data = {}
            kunci = []
            from apotek.models import Resep
            q = Resep.objects.select_related('kunjungan_pasien', 'nama_obat__nama_obat').filter(kunjungan_pasien__tgl_kunjungan__gte=request.POST.get("tanggal1"), kunjungan_pasien__tgl_kunjungan__lte=request.POST.get("tanggal2")).filter(nama_obat__nama_obat__is_okt=True)
            
            for object in q:
                kunci.append(object.nama_obat.nama_obat.nama_obat)
            kunci = list(set(kunci))
            kunci.sort()

            for item in kunci:
                r = q.filter(nama_obat__nama_obat__nama_obat=item)
                for data in r:                    
                    if item not in raw_data.keys():
                        raw_data[item] = [[
                            data.kunjungan_pasien.tgl_kunjungan.strftime('%Y-%m-%d'),
                            data.kunjungan_pasien.nama_pasien.nama_pasien,
                            "{} tahun".format(data.kunjungan_pasien.nama_pasien.umur()),
                            data.kunjungan_pasien.nama_pasien.alamat,
                            data.kunjungan_pasien.penulis_resep.nama_peresep,
                            data.jumlah,
                            data.nama_obat.nama_obat.satuan,
                        ]]
                    else:
                        raw_data[item].append([
                            data.kunjungan_pasien.tgl_kunjungan.strftime('%Y-%m-%d'),
                            data.kunjungan_pasien.nama_pasien.nama_pasien,
                            "{} tahun".format(data.kunjungan_pasien.nama_pasien.umur()),
                            data.kunjungan_pasien.nama_pasien.alamat,
                            data.kunjungan_pasien.penulis_resep.nama_peresep,
                            data.jumlah,
                            data.nama_obat.nama_obat.satuan,
                        ])
                        
        ctx = {
            'data': raw_data,
            'startdate': request.POST.get("tanggal1"),
            'enddate': request.POST.get("tanggal2"),
        }
        #return HttpResponse(json.dumps(raw_data), content_type="application/json")
        return render(request, 'laporan/lap_narko_psiko.html', context=ctx)
    else:
        form = TglForm()
        return render(request, 'laporan/form_lap_narko_psiko.html', {'form': form})

def tengok_stok_alkes(request):
    from apotek.models import StokObatGudang
    q = StokObatGudang.objects.select_related('nama_obat').filter(nama_obat__is_alkes=True)
    alkes = {}
    for item in q:
        if item.jml > 0:
            alkes[item.nama_obat.nama_obat] = [item.nama_obat.satuan, "ada"]
        else:
            alkes[item.nama_obat.nama_obat] = [item.nama_obat.satuan, "kosong"]

    alkes = OrderedDict(sorted(alkes.items()))
    ctx = {
        'data': alkes,
    }
    return render(request, 'laporan/tengok_stok_alkes.html', context=ctx)

def tengok_stok_obat(request):
    from apotek.models import StokObatGudang, StokObatApotek

    okt = {}
    q = StokObatGudang.objects.select_related('nama_obat').filter(nama_obat__is_okt=True)
    for item in q:
        okt[item.nama_obat.nama_obat] = [item.nama_obat.satuan, item.jml]
    q = StokObatApotek.objects.select_related('nama_obat').filter(nama_obat__is_okt=True)
    for item in q:
        if item.nama_obat.nama_obat not in okt:
            okt[item.nama_obat.nama_obat] = [item.nama_obat.satuan, item.jml]
        else:
            okt[item.nama_obat.nama_obat][1] += item.jml
    okt = OrderedDict(sorted(okt.items()))

    ab = {}
    q = StokObatGudang.objects.select_related('nama_obat').filter(nama_obat__is_ab=True)
    for item in q:
        ab[item.nama_obat.nama_obat] = [item.nama_obat.satuan, item.jml]
    q = StokObatApotek.objects.select_related('nama_obat').filter(nama_obat__is_ab=True)
    for item in q:
        if item.nama_obat.nama_obat not in ab:
            ab[item.nama_obat.nama_obat] = [item.nama_obat.satuan, item.jml]
        else:
            ab[item.nama_obat.nama_obat][1] += item.jml
    ab = OrderedDict(sorted(ab.items()))

    obt = {}
    q = StokObatGudang.objects.select_related('nama_obat').filter(nama_obat__is_alkes=False, nama_obat__is_ab=False, nama_obat__is_okt=False)
    for item in q:
        obt[item.nama_obat.nama_obat] = [item.nama_obat.satuan, item.jml]
    q = StokObatApotek.objects.select_related('nama_obat').filter(nama_obat__is_ab=True)
    for item in q:
        if item.nama_obat.nama_obat not in obt:
            obt[item.nama_obat.nama_obat] = [item.nama_obat.satuan, item.jml]
        else:
            obt[item.nama_obat.nama_obat][1] += item.jml
    obt = OrderedDict(sorted(obt.items()))
            
    ctx = {
        'okt': okt,
        'antibiotik': ab,
        'obat': obt,
    }
    return render(request, 'laporan/tengok_stok_obat.html', context=ctx)

@login_required
def so_apotek(request):
    from .forms import SingleTgl
    from apotek.models import StokObatApotek, KartuStokApotek, SOApotek

    stok_apotek = StokObatApotek.objects.all()
    form = SingleTgl(request.POST)
    context = {
        'data': stok_apotek
    }
    context['form'] = form
    json_data = {}
    
    if request.method == 'POST':
        if form.is_valid():
            
            #print(form['tanggal'].value())
            
            for obat in stok_apotek:
                #Ubah jumlah stok sesuai value yg diinput

                if int(request.POST[obat.nama_obat.nama_obat]) != obat.jml:
                    obat.jml = request.POST[obat.nama_obat.nama_obat]
                    obat.save()
                    json_data[obat.nama_obat.nama_obat] = [obat.jml, obat.jml]

                if int(request.POST[obat.nama_obat.nama_obat]) == obat.jml:
                    obat.jml = request.POST[obat.nama_obat.nama_obat]
                    obat.save()
                    json_data[obat.nama_obat.nama_obat] = [obat.jml, obat.jml]

                #Penyesuaian kartu stok tiap item
                query_kartu_apt = KartuStokApotek.objects.filter(nama_obat__nama_obat=obat)
                stok_apt_sebelum = query_kartu_apt[len(query_kartu_apt)-1]

                #Bila stok fisik lebih kecil daripada stok tercatat sebelumnya
                if int(request.POST[obat.nama_obat.nama_obat]) < stok_apt_sebelum.sisa_stok:
                    selisih = stok_apt_sebelum.sisa_stok - int(request.POST[obat.nama_obat.nama_obat])
                    #kartu_stok_apt_input = KartuStokApotek(nama_obat=obat.nama_obat, tgl=form['tanggal'].value(), unit="Staf", stok_keluar=selisih, sisa_stok=int(request.POST[obat.nama_obat.nama_obat]), ref=obat.id)
                    kartu_stok_apt_input = KartuStokApotek(nama_obat=obat.nama_obat, tgl=form['tanggal'].value(), unit="SO", stok_keluar=selisih, sisa_stok=int(request.POST[obat.nama_obat.nama_obat]), ref=obat.id)
                    kartu_stok_apt_input.save()
                    json_data[obat.nama_obat.nama_obat] = [stok_apt_sebelum.sisa_stok, int(request.POST[obat.nama_obat.nama_obat])]

                #Bila stok fisik lebih besar daripada stok tercatat sebelumnya
                if int(request.POST[obat.nama_obat.nama_obat]) > stok_apt_sebelum.sisa_stok:
                    selisih = int(request.POST[obat.nama_obat.nama_obat]) - stok_apt_sebelum.sisa_stok
                    #kartu_stok_apt_input = KartuStokApotek(nama_obat=obat.nama_obat, tgl=form['tanggal'].value(), unit="Retur", stok_terima=selisih, sisa_stok=int(request.POST[obat.nama_obat.nama_obat]), ref=obat.id)
                    kartu_stok_apt_input = KartuStokApotek(nama_obat=obat.nama_obat, tgl=form['tanggal'].value(), unit="SO", stok_keluar=selisih, sisa_stok=int(request.POST[obat.nama_obat.nama_obat]), ref=obat.id)
                    kartu_stok_apt_input.save()
                    json_data[obat.nama_obat.nama_obat] = [stok_apt_sebelum.sisa_stok, int(request.POST[obat.nama_obat.nama_obat])]

            data_so = SOApotek(tgl=form['tanggal'].value(), data=json_data)
            data_so.save()
                    
            return render(request, 'laporan/so_apotek.html', context)
            
    else:
        return render(request, 'laporan/so_apotek.html', context)

@login_required
def so_gudang(request):
    from .forms import SingleTgl
    from apotek.models import StokObatGudang, KartuStokGudang, SOGudang

    stok_gudang = StokObatGudang.objects.all()
    form = SingleTgl(request.POST)
    context = {
        'data': stok_gudang
    }
    context['form'] = form
    json_data = {}
    
    if request.method == 'POST':
        if form.is_valid():
            
            #print(form['tanggal'].value())
            
            for obat in stok_gudang:
                #Ubah jumlah stok sesuai value yg diinput

                if int(request.POST[obat.nama_obat.nama_obat]) != obat.jml:
                    obat.jml = request.POST[obat.nama_obat.nama_obat]
                    obat.save()
                    json_data[obat.nama_obat.nama_obat] = [obat.jml, obat.jml]

                if int(request.POST[obat.nama_obat.nama_obat]) == obat.jml:
                    obat.jml = request.POST[obat.nama_obat.nama_obat]
                    obat.save()
                    json_data[obat.nama_obat.nama_obat] = [obat.jml, obat.jml]

                #Penyesuaian kartu stok tiap item
                stok_gd_sebelum = KartuStokGudang.objects.filter(nama_obat__nama_obat=obat.nama_obat).last()
                try:
                    #Bila stok fisik lebih kecil daripada stok tercatat sebelumnya
                    if int(request.POST[obat.nama_obat.nama_obat]) < stok_gd_sebelum.sisa_stok:
                        selisih = stok_gd_sebelum.sisa_stok - int(request.POST[obat.nama_obat.nama_obat])
                        kartu_stok_gd_input = KartuStokGudang(nama_obat=obat.nama_obat, tgl=form['tanggal'].value(), unit="SO", stok_keluar=selisih, sisa_stok=int(request.POST[obat.nama_obat.nama_obat]))
                        kartu_stok_gd_input.save()
                        json_data[obat.nama_obat.nama_obat] = [stok_gd_sebelum.sisa_stok, int(request.POST[obat.nama_obat.nama_obat])]

                    #Bila stok fisik lebih besar daripada stok tercatat sebelumnya
                    if int(request.POST[obat.nama_obat.nama_obat]) > stok_gd_sebelum.sisa_stok:
                        selisih = int(request.POST[obat.nama_obat.nama_obat]) - stok_gd_sebelum.sisa_stok
                        kartu_stok_gd_input = KartuStokGudang(nama_obat=obat.nama_obat, tgl=form['tanggal'].value(), unit="SO", stok_terima=selisih, sisa_stok=int(request.POST[obat.nama_obat.nama_obat]))
                        kartu_stok_gd_input.save()
                        json_data[obat.nama_obat.nama_obat] = [stok_gd_sebelum.sisa_stok, int(request.POST[obat.nama_obat.nama_obat])]
                except AttributeError:
                    pass

            data_so = SOGudang(tgl=form['tanggal'].value(), data=json_data)
            data_so.save()
            return render(request, 'laporan/so_gudang.html', context)
            
    else:
        return render(request, 'laporan/so_gudang.html', context)

@login_required
def lap_generik(request):
    from .forms import GenerikForm
    
    if request.method == 'POST':
        from apotek.models import Resep
        from poli.models import DataKunjungan
        form = GenerikForm(request.POST)
        if form.is_valid():
            try:
                jml_lbr = DataKunjungan.objects.filter(tgl_kunjungan__gte=request.POST.get("tanggal1"), tgl_kunjungan__lte=request.POST.get("tanggal2")).filter(penulis_resep__nama_peresep=request.POST.get("pilihan")).count()

                tot_r = Resep.objects.filter(kunjungan_pasien__tgl_kunjungan__gte=request.POST.get("tanggal1"), kunjungan_pasien__tgl_kunjungan__lte=request.POST.get("tanggal2")).filter(kunjungan_pasien__penulis_resep__nama_peresep=request.POST.get("pilihan")).count()
                
                r_generik = Resep.objects.filter(kunjungan_pasien__tgl_kunjungan__gte=request.POST.get("tanggal1"), kunjungan_pasien__tgl_kunjungan__lte=request.POST.get("tanggal2")).filter(kunjungan_pasien__penulis_resep__nama_peresep=request.POST.get("pilihan")).filter(nama_obat__nama_obat__is_non_generik=True).count()
                
                ctx = {
                    'kalkulasi': {
                        'peresep': request.POST.get("pilihan"),
                        'jml_lembar': jml_lbr,
                        'total_resep': tot_r,
                        'resep_generik': tot_r - r_generik,
                        'persentase': "{0:.2f} %".format((tot_r - r_generik) / tot_r * 100),
                    },
                }
                ctx['form'] = form

                return render(request, 'laporan/form_lap_generik.html', context=ctx)

            except:
                return HttpResponse("Tidak ada data. Coba tanggal atau pilihan lain.")
    else:
        form = GenerikForm()
        
    return render(request, 'laporan/form_lap_generik.html', {'form': form})

@login_required
def lap_por(request):
    from .forms import PorForm
    if request.method == 'POST':
        from apotek.models import Resep
        form = PorForm(request.POST)
        if form.is_valid():
            try:
                #refdose = dict(Resep.ATURAN_PK)
                counter = 1
                raw_data = {}
                q = Resep.objects.select_related('kunjungan_pasien__nama_pasien', 'nama_obat__nama_obat').filter(kunjungan_pasien__tgl_kunjungan__gte=request.POST.get("tanggal1"), kunjungan_pasien__tgl_kunjungan__lte=request.POST.get("tanggal2")).filter(kunjungan_pasien__diagnosa__diagnosa=request.POST.get("pilihan")).order_by('kunjungan_pasien__tgl_kunjungan').iterator()
                
                for data in q:
                    if not raw_data:
                        raw_data[counter] = {"nama": data.kunjungan_pasien.nama_pasien.nama_pasien, "tgl": data.kunjungan_pasien.tgl_kunjungan.strftime("%Y-%m-%d"), "obat": [(data.nama_obat.nama_obat.nama_obat, data.aturan_pakai, data.lama_pengobatan, data.jumlah, data.nama_obat.nama_obat.is_ab)], "id": data.kunjungan_pasien.id, "usia": data.kunjungan_pasien.nama_pasien.umur()}
                    elif data.kunjungan_pasien.nama_pasien.nama_pasien == raw_data[counter]["nama"] and data.kunjungan_pasien.id == raw_data[counter]["id"]:
                        raw_data[counter]["obat"].append((data.nama_obat.nama_obat.nama_obat, data.aturan_pakai, data.lama_pengobatan, data.jumlah, data.nama_obat.nama_obat.is_ab))
                    else:
                        counter += 1
                        raw_data[counter] = {"nama": data.kunjungan_pasien.nama_pasien.nama_pasien, "tgl": data.kunjungan_pasien.tgl_kunjungan.strftime("%Y-%m-%d"), "obat": [(data.nama_obat.nama_obat.nama_obat, data.aturan_pakai, data.lama_pengobatan, data.jumlah, data.nama_obat.nama_obat.is_ab)], "id": data.kunjungan_pasien.id, "usia": data.kunjungan_pasien.nama_pasien.umur()}
                        
                ctx = {
                    'data': raw_data,
                    'input': {
                        'date1': request.POST.get("tanggal1"),
                        'date2': request.POST.get("tanggal2"),
                        'diag': request.POST.get("pilihan"),
                    }
                }
                return render(request, 'laporan/lap_por.html', context=ctx)

            except Exception as e:
                return HttpResponse(e)

    else:
        form = PorForm()
    return render(request, 'laporan/form_lap_por.html', {'form': form})

@login_required
def lap_pasien(request):
    from .forms import TglForm
    if request.method == 'POST':
        from apotek.models import Resep
        form = TglForm(request.POST)
        if form.is_valid():
            try:
                #refdose = dict(Resep.ATURAN_PK)
                counter = 1
                raw_data = {}
                q = Resep.objects.select_related('kunjungan_pasien__nama_pasien', 'nama_obat__nama_obat').filter(kunjungan_pasien__tgl_kunjungan__gte=request.POST.get("tanggal1"), kunjungan_pasien__tgl_kunjungan__lte=request.POST.get("tanggal2")).order_by('kunjungan_pasien__tgl_kunjungan').iterator()
                
                for data in q:
                    if not raw_data:
                        raw_data[counter] = {"nama": data.kunjungan_pasien.nama_pasien.nama_pasien, "tgl": data.kunjungan_pasien.tgl_kunjungan.strftime("%Y-%m-%d"), "obat": [(data.nama_obat.nama_obat.nama_obat, data.aturan_pakai, data.lama_pengobatan, data.jumlah, data.nama_obat.nama_obat.is_ab)], "id": data.kunjungan_pasien.id, "usia": data.kunjungan_pasien.nama_pasien.umur(), "diagn": [x.diagnosa for x in data.kunjungan_pasien.diagnosa.distinct()]}
                    elif data.kunjungan_pasien.nama_pasien.nama_pasien == raw_data[counter]["nama"] and data.kunjungan_pasien.id == raw_data[counter]["id"]:
                        raw_data[counter]["obat"].append((data.nama_obat.nama_obat.nama_obat, data.aturan_pakai, data.lama_pengobatan, data.jumlah, data.nama_obat.nama_obat.is_ab))
                    else:
                        counter += 1
                        raw_data[counter] = {"nama": data.kunjungan_pasien.nama_pasien.nama_pasien, "tgl": data.kunjungan_pasien.tgl_kunjungan.strftime("%Y-%m-%d"), "obat": [(data.nama_obat.nama_obat.nama_obat, data.aturan_pakai, data.lama_pengobatan, data.jumlah, data.nama_obat.nama_obat.is_ab)], "id": data.kunjungan_pasien.id, "usia": data.kunjungan_pasien.nama_pasien.umur(), "diagn": [x.diagnosa for x in data.kunjungan_pasien.diagnosa.distinct()]}
                        
                ctx = {
                    'data': raw_data,
                    'input': {
                        'date1': request.POST.get("tanggal1"),
                        'date2': request.POST.get("tanggal2"),
                    }
                }
                #a = [x.diagnosa for x in raw_data[1]["diagn"].distinct()]
                return render(request, 'laporan/lap_pasien.html', context=ctx)
                #return HttpResponse(ctx['data'])

            except Exception as e:
                return HttpResponse(e)

    else:
        form = TglForm()
    return render(request, 'laporan/form_lap_pasien.html', {'form': form})

@login_required
def manifest_distribusi(request):
    from apotek.models import Resep, Pengeluaran
    from .forms import TglForm
    import datetime
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = TglForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            raw_data_bmhp_apt = {}
            raw_data_bmhp_unit = {}
            kunci = []
            cleaned_data_bmhp_unit = {}
            
            q = Resep.objects.select_related('nama_obat__nama_obat').filter(kunjungan_pasien__tgl_kunjungan__gte=request.POST.get("tanggal1"), kunjungan_pasien__tgl_kunjungan__lte=request.POST.get("tanggal2")).iterator()
            q2 = Pengeluaran.objects.select_related('nama_barang__nama_obat').filter(keluar_barang__tgl_keluar__gte=request.POST.get("tanggal1"), keluar_barang__tgl_keluar__lte=request.POST.get("tanggal2")).exclude(keluar_barang__tujuan__nama="APOTEK").iterator()

            for data in q:
                if data.nama_obat.nama_obat.nama_obat not in raw_data_bmhp_apt:
                    raw_data_bmhp_apt[data.nama_obat.nama_obat.nama_obat] = data.jumlah
                else:
                    raw_data_bmhp_apt[data.nama_obat.nama_obat.nama_obat] += data.jumlah

            cleaned_data_obat = dict(sorted(raw_data_bmhp_apt.items(), key=operator.itemgetter(1),reverse=True))
            cleaned_data_obat = OrderedDict(sorted(cleaned_data_obat.items()))

            for data in q2:
                a = {}
                a[data.nama_barang.nama_obat.nama_obat] = data.jumlah
                if data.keluar_barang.tujuan.nama not in raw_data_bmhp_unit.keys():
                    raw_data_bmhp_unit[data.keluar_barang.tujuan.nama] = [a]
                    kunci.append(data.keluar_barang.tujuan.nama)
                else:                    
                    raw_data_bmhp_unit[data.keluar_barang.tujuan.nama].append(a)

            for d in kunci:
                c = Counter()
                for x in raw_data_bmhp_unit[d]:
                    c.update(x)
                    cleaned_data_bmhp_unit[d] = dict(c)
            
            context = {
                'startdate': datetime.datetime.strptime(request.POST.get("tanggal1"), "%Y-%m-%d").date(),
                'enddate': datetime.datetime.strptime(request.POST.get("tanggal2"), "%Y-%m-%d").date(),
                'val': cleaned_data_obat,
                'unit': cleaned_data_bmhp_unit,
            }
            return render(request, 'laporan/manifest_distribusi.html', context)

    # if a GET (or any other method) we'll create a blank form
    else:
        form = TglForm()

    return render(request, 'laporan/form_manifest_dist.html', {'form': form})
   
