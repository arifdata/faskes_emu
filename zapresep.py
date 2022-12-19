import requests
from openpyxl import load_workbook
import re
import sys

EXCEL_FILE = "Laporan Harian - Pelayanan Pasien.xlsx"
try:
    wb = load_workbook(filename=EXCEL_FILE)
except FileNotFoundError:
    sys.exit('Error: File excel harus bernama "STS.xlsx".')
sr = wb.active

def getObat(obatraw):
    items = []
    counter = 0
    rawText = obatraw
    nama_bmhp = re.findall(r'-.*', rawText)
    for i in nama_bmhp:
        items.append([i.removeprefix('- ')])

    jml = re.findall(r'Jumlah.*', rawText)
    for i in jml:
        items[counter].append(int(i.removeprefix('Jumlah : ')))
        counter += 1
    counter = 0

    signa = re.findall(r'Signa.*', rawText)
    for i in signa:
        items[counter].append(i.removeprefix('Signa : '))
        counter += 1
    counter = 0

    for i in items:
        try:
            x = i[1] / int(i[2][0])
            if int(x) == 0:
                items[counter].append(3)
                counter += 1
            else:
                items[counter].append(int(x))
                counter += 1
        except ValueError:
            items[counter].append(3)
            counter += 1
    counter = 0

    return items

def getIDPasien(sesi, nama_pasien, no_kartu, usia, alamat, no_hp):
    url = "http://localhost:8000/api/datapasien/?format=json&search={} {}".format(nama_pasien, no_kartu)
    q = sesi.get(url)
    if q.json()["count"] == 0:
        data = {
            "nama_pasien": nama_pasien,
            "no_kartu": no_kartu,
            "usia": usia,
            "alamat": alamat,
            "no_hp": no_hp
        }
        q = sesi.post("http://localhost:8000/api/datapasien/", json=data)
        url = "http://localhost:8000/api/datapasien/?format=json&search={} {}".format(nama_pasien, no_kartu)
        q = sesi.get(url)
        for pas in q.json()["results"]:
            if nama_pasien == pas["nama_pasien"]:
                return pas["id"]
    else:
        for pas in q.json()["results"]:
            if nama_pasien == pas["nama_pasien"]:
                return pas["id"]

def getIDPeresep(sesi, penulis):
    url = "http://localhost:8000/api/peresep/?format=json&search={}".format(penulis)
    q = sesi.get(url)
    for dr in q.json():
        if penulis == dr["nama_peresep"]:
            return dr["id"]

def getIDObat(sesi, obat):
    url = "http://localhost:8000/api/stokobatapotek/?format=json&search={}".format(obat)
    q = sesi.get(url)
    for o in q.json()["results"]:
        if obat == o["nama_obat"]:
            return o["id"]

def getIDDiagnosa(sesi, rowNum):
    diag = []
    idDiag = []
    diag.append(sr.cell(row=rowNum, column=64).value)
    diag.append(sr.cell(row=rowNum, column=66).value)
    diag.append(sr.cell(row=rowNum, column=68).value)
    diag.append(sr.cell(row=rowNum, column=70).value)
    diag.append(sr.cell(row=rowNum, column=72).value)
    
    diag = [i for i in diag if i is not None]

    for namapenyakit in diag:
        url = "http://localhost:8000/api/diagnosa/?format=json&search={}".format(namapenyakit)
        q = sesi.get(url)
        for res in q.json():
            if namapenyakit == res["diagnosa"]:
                idDiag.append(res["id"])
    return idDiag

def coba_post(sesi):
    print("======= Bersiap input resep =======")
    r = sesi.get("http://localhost:8000/app/poli/datakunjungan/add/")
    rawText = r.text
    regex = re.search(r'csrfmiddlewaretoken" value=".*">', rawText)
    token = regex.group().removeprefix('csrfmiddlewaretoken" value="').removesuffix('">')
    for rowNum in range(26, sr.max_row + 1):
        if sr.cell(row=rowNum, column=75).value != None:
            files = {
                "csrfmiddlewaretoken": token,
                "diagnosa": [],
                "resep_set-TOTAL_FORMS": 4,
                "resep_set-INITIAL_FORMS": 0,
                "_save": "Simpan"
            }
            
            c = 0
            nama_pasien = sr.cell(row=rowNum, column=3).value
            no_kartu = sr.cell(row=rowNum, column=5).value if sr.cell(row=rowNum, column=5).value != None else "belum_ada"
            usia = sr.cell(row=rowNum, column=18).value
            alamat = sr.cell(row=rowNum, column=16).value
            no_hp = sr.cell(row=rowNum, column=10).value
            idPasien = getIDPasien(sesi, nama_pasien, no_kartu, usia, alamat, no_hp)
            idPeresep = getIDPeresep(sesi, sr.cell(row=rowNum, column=29).value)

            tmpNamaObat = "resep_set-{}-nama_obat"
            tmpJmlObat = "resep_set-{}-jumlah"
            tmpDosisObat = "resep_set-{}-aturan_pakai"
            tmpLamaMinumObat = "resep_set-{}-lama_pengobatan"

            listObat = getObat(sr.cell(row=rowNum, column=75).value)
            for obat in listObat:
                idObat = getIDObat(sesi, obat[0])
                files[tmpNamaObat.format(c)] = idObat
                files[tmpJmlObat.format(c)] = obat[1]
                files[tmpDosisObat.format(c)] = obat[2]
                files[tmpLamaMinumObat.format(c)] = obat[3]
                c += 1

            listIDDiagnosa = getIDDiagnosa(sesi, rowNum)
            for diagnosa in listIDDiagnosa:
                files["diagnosa"].append(diagnosa)

            files["nama_pasien"] = idPasien
            files["tgl_kunjungan"] = sr.cell(row=rowNum, column=15).value[0:10]
            files["penulis_resep"] = idPeresep

            r = sesi.post("http://localhost:8000/app/poli/datakunjungan/add/", data=files)
            if "Stok di apotek tidak mencukupi." in r.text:
                msg = "{} gagal input karena stok obat berikut tidak cukup:".format(sr.cell(row=rowNum, column=3).value)
                print(msg)
                html = r.text
                stokkurang = re.findall(r'<ul class="errorlist"><li>Stok di apotek tidak mencukupi.*\n.*">', html)
                for regex in stokkurang:
                    regexItem = re.search(r'id_resep_set-\d-jumlah', regex)
                    idItem = int(regexItem.group().removeprefix("id_resep_set-").removesuffix("-jumlah"))
                    tpl = '<select name="resep_set-{}-nama_obat" id="id_resep_set-{}-nama_obat".*\n.*option>'.format(idItem, idItem)
                    itemRegex = re.search(tpl, html)
                    itemStok = itemRegex.group().split("selected>")[1].removesuffix("</option>")
                    print("{} - [{}]".format(itemStok, obat[1]))
                print("")
            else:
                msg = "Berhasil input {}.".format(sr.cell(row=rowNum, column=3).value)
                print(msg)
                print("")
        
def cekDiagnosa(sesi, diag):
    r = sesi.get("http://localhost:8000/api/diagnosa/")
    q = r.json()
    c = 0
    daftar_diagnosa = [d["diagnosa"] for d in q]

    for diagnosa in diag:
        if diagnosa not in daftar_diagnosa:
            json = {
                "diagnosa" : diagnosa
            }
            r = sesi.post("http://localhost:8000/api/diagnosa/", json=json)
            print("Menginput diagnosa: {}".format(diagnosa))
            c += 1
        else:
            pass
    print("Cek diagnosa selesai. Input {} diagnosa baru.".format(c))

def getDiagnosa():
    diag = []
    for rowNum in range(26, sr.max_row + 1):
        if sr.cell(row=rowNum, column=75).value == None:
            pass
        else:
            diag.append(sr.cell(row=rowNum, column=64).value)
            diag.append(sr.cell(row=rowNum, column=66).value)
            diag.append(sr.cell(row=rowNum, column=68).value)
            diag.append(sr.cell(row=rowNum, column=70).value)
            diag.append(sr.cell(row=rowNum, column=72).value)
    
    diag = [i for i in diag if i is not None]
    return diag

def cekPeresep(sesi, peresep):
    r = sesi.get("http://localhost:8000/api/peresep/")
    q = r.json()
    c = 0
    daftar_peresep = [r["nama_peresep"] for r in q]
    for penulis in peresep:
        if penulis not in daftar_peresep:
            json = {
                "nama_peresep": penulis
            }
            r = sesi.post("http://localhost:8000/api/peresep/", json=json)
            print("Menginput peresep baru: {}".format(penulis))
            c += 1
        else:
            pass
    print("Cek peresep selesai. Input {} peresep baru.".format(c))
            

def getPeresep():
    peresep = []
    for rowNum in range(26, sr.max_row + 1):
        if sr.cell(row=rowNum, column=75).value == None:
            pass
        else:
            peresep.append(sr.cell(row=rowNum, column=29).value)
    peresep = list(set(peresep))
    return peresep

def main():
    with requests.Session() as sesi:

        #pasien = getPasien(sesi)
    
        diag = getDiagnosa()
        cekDiagnosa(sesi, diag)

        peresep = getPeresep()
        cekPeresep(sesi, peresep)

        r = sesi.get("http://localhost:8000/app/login/")
        rawText = r.text
        regex = re.search(r'csrfmiddlewaretoken" value=".*">', rawText)
        token = regex.group().removeprefix('csrfmiddlewaretoken" value="').removesuffix('">')

        login_data = {
            'csrfmiddlewaretoken': token,
            'username': 'admin',
            'password': 'puskesmas'
        }
        r = sesi.post("http://localhost:8000/app/login/", data=login_data)

        coba_post(sesi)
    
if __name__ == "__main__":
    main()
